import os
import tempfile
from datetime import datetime, timezone
from src.celery_app import app


@app.task(bind=True, max_retries=3)
def generate_batch_report(
    self,
    batch_id: int,
    format: str = "excel",
    user_email: str = None,
):
    """
    Генерация Excel отчёта по партии.
    
    Создаёт файл с тремя листами:
    - Информация о партии
    - Список продукции
    - Статистика
    
    Загружает файл в MinIO и возвращает presigned URL.
    """
    import asyncio
    from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    from src.data.models.batch import Batch
    from src.core.config import settings
    from src.storage.minio_service import minio_service
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    async def get_batch_data():
        engine = create_async_engine(settings.database_url)
        SessionLocal = async_sessionmaker(engine, class_=AsyncSession)

        async with SessionLocal() as db:
            result = await db.execute(
                select(Batch)
                .options(selectinload(Batch.products))
                .where(Batch.id == batch_id)
            )
            batch = result.scalar_one_or_none()

        await engine.dispose()
        return batch

    try:
        batch = asyncio.run(get_batch_data())
        if not batch:
            return {"success": False, "error": f"Batch {batch_id} not found"}

        # Создаём Excel файл
        wb = openpyxl.Workbook()

        # ===== Лист 1: Информация о партии =====
        ws1 = wb.active
        ws1.title = "Информация о партии"

        # Стиль для заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="366092")

        info_data = [
            ("Номер партии", batch.batch_number),
            ("Дата партии", str(batch.batch_date)),
            ("Статус", "Закрыта" if batch.is_closed else "Открыта"),
            ("Смена", batch.shift),
            ("Бригада", batch.team),
            ("Номенклатура", batch.nomenclature),
            ("Код ЕКН", batch.ekn_code),
            ("Начало смены", str(batch.shift_start)),
            ("Окончание смены", str(batch.shift_end)),
        ]

        for row_idx, (label, value) in enumerate(info_data, 1):
            ws1.cell(row=row_idx, column=1, value=label).font = header_font
            ws1.cell(row=row_idx, column=2, value=str(value))

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 35

        # ===== Лист 2: Продукция =====
        ws2 = wb.create_sheet("Продукция")
        headers = ["ID", "Уникальный код", "Агрегирована", "Дата агрегации"]

        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")

        for row_idx, product in enumerate(batch.products, 2):
            ws2.cell(row=row_idx, column=1, value=product.id)
            ws2.cell(row=row_idx, column=2, value=product.unique_code)
            ws2.cell(row=row_idx, column=3, value="Да" if product.is_aggregated else "Нет")
            ws2.cell(
                row=row_idx, column=4,
                value=str(product.aggregated_at) if product.aggregated_at else "-"
            )

        for col in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col].width = 25

        # ===== Лист 3: Статистика =====
        ws3 = wb.create_sheet("Статистика")
        total = len(batch.products)
        aggregated = sum(1 for p in batch.products if p.is_aggregated)
        remaining = total - aggregated
        rate = round(aggregated / total * 100, 2) if total > 0 else 0

        stats = [
            ("Всего продукции", total),
            ("Агрегировано", aggregated),
            ("Осталось", remaining),
            ("Процент выполнения", f"{rate}%"),
        ]

        for row_idx, (label, value) in enumerate(stats, 1):
            ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=row_idx, column=2, value=str(value))

        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 20

        # Сохраняем файл во временную папку
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False, prefix=f"batch_{batch_id}_"
        ) as tmp:
            tmp_path = tmp.name

        wb.save(tmp_path)
        file_size = os.path.getsize(tmp_path)
        object_name = f"batch_{batch_id}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Загружаем в MinIO и получаем ссылку
        url = minio_service.upload_file(
            bucket="reports",
            file_path=tmp_path,
            object_name=object_name,
            expires_days=7,
        )

        # Удаляем временный файл
        os.unlink(tmp_path)

        return {
            "success": True,
            "file_url": url,
            "file_name": object_name,
            "file_size": file_size,
            "expires_at": datetime.now(timezone.utc).isoformat(),
        }

    except Exception as exc:
        raise self.retry(exc=exc, countdown=60)
